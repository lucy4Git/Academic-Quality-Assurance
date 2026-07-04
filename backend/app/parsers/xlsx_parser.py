"""XLSX parser using openpyxl.

Iterates every sheet, reads all non-empty cells as text, and builds a
flat text representation suitable for keyword classification.  Sheet names,
total row/column counts, and a sample of the first sheet's headers (first row)
are stored in metadata.
"""

from __future__ import annotations

import io

from app.parsers.base import DocumentParser, ExtractionResult, ParserError


class XlsxParser(DocumentParser):
    """Extract cell text from XLSX files."""

    @property
    def supported_mime_types(self) -> frozenset[str]:
        return frozenset(
            {
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            }
        )

    async def extract(self, content: bytes, filename: str) -> ExtractionResult:
        return await self._run_sync(self._extract_sync, content, filename)

    @staticmethod
    def _extract_sync(content: bytes, filename: str) -> ExtractionResult:
        try:
            import openpyxl  # noqa: PLC0415
        except ImportError:
            raise ParserError(
                "openpyxl is not installed. Run: pip install openpyxl",
                recoverable=False,
            )

        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(content),
                read_only=True,
                data_only=True,
            )
        except Exception as exc:
            raise ParserError(f"Cannot open XLSX '{filename}': {exc}") from exc

        sheet_names = wb.sheetnames
        parts: list[str] = []
        total_rows = 0
        total_cols = 0
        first_headers: list[str] = []

        for sheet_idx, sheet_name in enumerate(sheet_names):
            ws = wb[sheet_name]
            sheet_rows = 0
            sheet_cols = 0
            is_first = sheet_idx == 0

            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                non_empty = [str(cell) for cell in row if cell is not None]
                if not non_empty:
                    continue
                sheet_rows += 1
                sheet_cols = max(sheet_cols, len(non_empty))
                if is_first and row_idx == 0:
                    first_headers = non_empty
                parts.extend(non_empty)

            total_rows += sheet_rows
            total_cols = max(total_cols, sheet_cols)

        wb.close()

        full_text = " | ".join(parts).strip()

        return ExtractionResult(
            text=full_text,
            word_count=len(full_text.split()),
            parser_name="xlsx",
            page_count=len(sheet_names),  # "pages" ≈ sheets
            metadata={
                "sheet_names": sheet_names,
                "sheet_count": len(sheet_names),
                "row_count": total_rows,
                "col_count": total_cols,
                "first_sheet_headers": first_headers,
            },
        )
