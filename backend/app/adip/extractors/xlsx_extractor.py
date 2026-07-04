"""ADIP XLSX / CSV extractor using openpyxl."""

from __future__ import annotations

import csv
import io
from pathlib import Path

from app.adip.extractors.base import (
    BaseExtractor,
    DocumentMetadata,
    ExtractedChunk,
    ExtractedTable,
    ExtractionResult,
)

XLSX_MIMES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/vnd.ms-excel",
}
CSV_MIMES = {"text/csv", "text/plain"}
METHOD_XLSX = "openpyxl"
METHOD_CSV = "csv_builtin"


class XLSXExtractor(BaseExtractor):
    """Extracts tables and text from XLSX spreadsheets."""

    def can_handle(self, mime_type: str, file_path: Path) -> bool:
        return mime_type in XLSX_MIMES or str(file_path).lower().endswith((".xlsx", ".xls"))

    def extract(self, file_path: Path) -> ExtractionResult:
        try:
            import openpyxl
        except ImportError:
            return ExtractionResult(
                extraction_method=METHOD_XLSX,
                error="openpyxl not installed. Run: pip install openpyxl",
            )

        chunks: list[ExtractedChunk] = []
        tables: list[ExtractedTable] = []
        warnings: list[str] = []
        sequence = 0

        try:
            wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        except Exception as exc:
            return ExtractionResult(extraction_method=METHOD_XLSX, error=f"Cannot open XLSX: {exc}")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # Find first non-empty row as header
            header_row_idx = next(
                (i for i, row in enumerate(rows) if any(c is not None for c in row)), 0
            )
            headers = [str(c).strip() if c is not None else f"Column_{j}" for j, c in enumerate(rows[header_row_idx])]

            data_rows = []
            for row in rows[header_row_idx + 1:]:
                if not any(c is not None for c in row):
                    continue
                cells = [str(c).strip() if c is not None else "" for c in row]
                data_rows.append({headers[i]: cells[i] if i < len(cells) else "" for i in range(len(headers))})
                # Also produce text chunks for full-text indexing
                row_text = " | ".join(f"{headers[i]}: {cells[i]}" for i in range(len(headers)) if i < len(cells) and cells[i])
                if row_text.strip():
                    chunks.append(ExtractedChunk(
                        text=row_text,
                        chunk_type="cell_value",
                        sheet_name=sheet_name,
                        extraction_method=METHOD_XLSX,
                        sequence_index=sequence,
                    ))
                    sequence += 1

            tables.append(ExtractedTable(
                page_number=None,
                table_index=len(tables),
                header_row=headers,
                data_rows=data_rows,
                extraction_method=METHOD_XLSX,
                sheet_name=sheet_name,
            ))

        wb.close()
        return ExtractionResult(
            extraction_method=METHOD_XLSX,
            chunks=chunks,
            tables=tables,
            metadata=DocumentMetadata(),
            extraction_quality=0.93,
            warnings=warnings,
        )


class CSVExtractor(BaseExtractor):
    """Extracts tables from CSV files."""

    def can_handle(self, mime_type: str, file_path: Path) -> bool:
        return mime_type in CSV_MIMES or str(file_path).lower().endswith(".csv")

    def extract(self, file_path: Path) -> ExtractionResult:
        chunks: list[ExtractedChunk] = []
        sequence = 0
        try:
            content = file_path.read_text(encoding="utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(content))
            headers = reader.fieldnames or []
            data_rows = []
            for row in reader:
                data_rows.append(dict(row))
                row_text = " | ".join(f"{k}: {v}" for k, v in row.items() if v)
                if row_text.strip():
                    chunks.append(ExtractedChunk(
                        text=row_text,
                        chunk_type="cell_value",
                        extraction_method=METHOD_CSV,
                        sequence_index=sequence,
                    ))
                    sequence += 1
            tables = [ExtractedTable(
                page_number=None,
                table_index=0,
                header_row=list(headers),
                data_rows=data_rows,
                extraction_method=METHOD_CSV,
            )]
            return ExtractionResult(
                extraction_method=METHOD_CSV,
                chunks=chunks,
                tables=tables,
                metadata=DocumentMetadata(),
                extraction_quality=0.95,
            )
        except Exception as exc:
            return ExtractionResult(extraction_method=METHOD_CSV, error=f"CSV error: {exc}")
