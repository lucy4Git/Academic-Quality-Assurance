"""Reporting export service — CSV, Excel, and PDF placeholder.

All functions return bytes ready to be streamed as a FastAPI response.

PDF note
--------
`reportlab` is not in the dependency tree. The `export_pdf` function returns
the report content as UTF-8 encoded bytes with Content-Type text/plain and a
clear notice that full PDF rendering is a planned feature.  This satisfies the
sprint requirement of a "documented placeholder".
"""

from __future__ import annotations

import csv
import io
from datetime import datetime, timezone
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# CSV export
# ---------------------------------------------------------------------------


def export_csv(rows: list[dict[str, Any]], title: str = "AQAA Report") -> bytes:
    """Serialise a list of dicts to CSV bytes (UTF-8 with BOM for Excel compat)."""
    if not rows:
        return b"\xef\xbb\xbfNo data available.\n"

    buf = io.StringIO()
    fieldnames = list(rows[0].keys())
    writer = csv.DictWriter(buf, fieldnames=fieldnames, extrasaction="ignore")
    writer.writeheader()
    writer.writerows(rows)

    content = buf.getvalue()
    return ("﻿" + content).encode("utf-8")  # UTF-8 BOM (U+FEFF)


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------


def export_excel(
    rows: list[dict[str, Any]],
    sheet_name: str = "Report",
    title: str = "AQAA Report",
) -> bytes:
    """Serialise a list of dicts to an Excel workbook (xlsx) bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]  # Excel sheet name limit

    # Header row
    if not rows:
        ws.append(["No data available."])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    headers = list(rows[0].keys())
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=str(header).replace("_", " ").title())
        cell.fill = header_fill
        cell.font = header_font

    # Data rows
    for row in rows:
        row_vals = []
        for h in headers:
            val = row.get(h)
            if isinstance(val, datetime):
                row_vals.append(val.strftime("%Y-%m-%d %H:%M UTC"))
            elif isinstance(val, list):
                row_vals.append(", ".join(str(v) for v in val))
            else:
                row_vals.append(val)
        ws.append(row_vals)

    # Auto-size columns (rough)
    for col_idx in range(1, len(headers) + 1):
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 20

    # Metadata sheet
    meta_ws = wb.create_sheet("Metadata")
    meta_ws.append(["Report", title])
    meta_ws.append(["Generated at", datetime.now(tz=timezone.utc).strftime("%Y-%m-%d %H:%M UTC")])
    meta_ws.append(["Platform", "AQAA — Academic Quality Assurance Agent"])
    meta_ws.append(["Row count", len(rows)])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# PDF placeholder
# ---------------------------------------------------------------------------

_PDF_PLACEHOLDER_HEADER = """\
================================================================================
AQAA — Academic Quality Assurance Agent
PDF Export Placeholder
================================================================================

NOTE: Full PDF generation (reportlab-based) is planned but not yet enabled in
      this deployment.  This file contains the report in plain-text format.

      To generate a proper PDF, configure the `reportlab` package and implement
      the `export_pdf_full` function in reporting/export_service.py.

================================================================================

"""


def export_pdf_placeholder(report_lines: list[str], title: str = "AQAA Report") -> bytes:
    """Return a plain-text report as bytes (PDF generation not yet implemented)."""
    body = "\n".join(report_lines)
    full_text = _PDF_PLACEHOLDER_HEADER + f"REPORT: {title}\n\n" + body + "\n"
    return full_text.encode("utf-8")


def build_dashboard_report_lines(dashboard_data: dict[str, Any]) -> list[str]:
    """Build a plain-text report body from a dashboard dict."""
    lines: list[str] = [
        f"Generated at: {dashboard_data.get('generated_at', 'N/A')}",
        "",
        "PLATFORM SUMMARY",
        "-" * 40,
        f"Institutions:  {dashboard_data.get('institution_count', 0)}",
        f"Faculties:     {dashboard_data.get('faculty_count', 0)}",
        f"Departments:   {dashboard_data.get('department_count', 0)}",
        f"Programmes:    {dashboard_data.get('programme_count', 0)}",
        f"Modules:       {dashboard_data.get('module_count', 0)}",
        f"Audit runs:    {dashboard_data.get('audit_run_count', 0)}",
        f"  Completed:   {dashboard_data.get('completed_audit_count', 0)}",
        f"  Failed:      {dashboard_data.get('failed_audit_count', 0)}",
        f"Evidence files:{dashboard_data.get('evidence_file_count', 0)}",
        "",
        "KNOWLEDGE INDEX STATUS",
        "-" * 40,
    ]
    for entry in dashboard_data.get("knowledge_index_status", []):
        status = "INDEXED" if entry.get("indexed") else "NOT INDEXED"
        lines.append(
            f"  {entry.get('institution_code')} {entry.get('ikp_version')} "
            f"({entry.get('collection')}): {status}"
        )
    lines.append("")
    lines.append("PER-INSTITUTION BREAKDOWN")
    lines.append("-" * 40)
    for inst in dashboard_data.get("by_institution", []):
        lines.append(f"  {inst.get('institution_code')} — {inst.get('institution_name')}")
        lines.append(f"    Faculties: {inst.get('faculty_count')}  |  Departments: {inst.get('department_count')}")
        lines.append(f"    Programmes: {inst.get('programme_count')}  |  Modules: {inst.get('module_count')}")
        lines.append(f"    Audit runs: {inst.get('audit_run_count')}  |  Files: {inst.get('evidence_file_count')}")
        qs = "Indexed" if inst.get("knowledge_indexed") else "Not indexed"
        lines.append(f"    Qdrant: {qs}")
        lines.append("")
    return lines
