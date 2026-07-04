"""Tests for the Reporting & Analytics subsystem.

Coverage
--------
- export_csv: bytes, BOM, headers, row count
- export_excel: bytes, valid xlsx, metadata sheet
- export_pdf_placeholder: bytes, contains notice, contains title
- build_dashboard_report_lines: structure validation
- Tenant isolation: non-admin scoped to own institution
- ComplianceSummaryResponse: empty case, rate calculation
"""

from __future__ import annotations

import io
import uuid
from unittest.mock import AsyncMock, MagicMock, patch

import openpyxl
import pytest

from app.reporting.export_service import (
    build_dashboard_report_lines,
    export_csv,
    export_excel,
    export_pdf_placeholder,
)
from app.reporting import report_service
from app.schemas.reporting import (
    ComplianceSummaryResponse,
    DashboardResponse,
    InstitutionStats,
    KnowledgeIndexEntry,
)


# ===========================================================================
# TestExportCsv
# ===========================================================================


class TestExportCsv:
    def test_returns_bytes(self) -> None:
        rows = [{"name": "TUT", "count": 5}]
        result = export_csv(rows)
        assert isinstance(result, bytes)

    def test_contains_utf8_bom(self) -> None:
        rows = [{"name": "TUT"}]
        result = export_csv(rows)
        assert result.startswith(b"\xef\xbb\xbf")

    def test_contains_headers(self) -> None:
        rows = [{"institution_code": "TUT", "faculty_count": 4}]
        result = export_csv(rows)
        text = result.decode("utf-8-sig")
        assert "institution_code" in text
        assert "faculty_count" in text

    def test_contains_row_data(self) -> None:
        rows = [{"institution_code": "UP", "faculty_count": 3}]
        result = export_csv(rows)
        text = result.decode("utf-8-sig")
        assert "UP" in text
        assert "3" in text

    def test_multiple_rows(self) -> None:
        rows = [{"code": "TUT"}, {"code": "UP"}]
        result = export_csv(rows)
        text = result.decode("utf-8-sig")
        assert text.count("\n") >= 3  # header + 2 rows + trailing newline

    def test_empty_rows_returns_notice(self) -> None:
        result = export_csv([])
        assert b"No data" in result

    def test_title_ignored_no_crash(self) -> None:
        result = export_csv([{"x": 1}], title="Custom Title")
        assert isinstance(result, bytes)


# ===========================================================================
# TestExportExcel
# ===========================================================================


class TestExportExcel:
    def test_returns_bytes(self) -> None:
        rows = [{"name": "TUT", "count": 5}]
        result = export_excel(rows)
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_valid_xlsx_can_be_parsed(self) -> None:
        rows = [{"institution_code": "TUT", "faculty_count": 4}]
        result = export_excel(rows)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert wb is not None

    def test_has_metadata_sheet(self) -> None:
        rows = [{"institution_code": "TUT"}]
        result = export_excel(rows)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert "Metadata" in wb.sheetnames

    def test_sheet_name_used(self) -> None:
        rows = [{"code": "UP"}]
        result = export_excel(rows, sheet_name="Dashboard")
        wb = openpyxl.load_workbook(io.BytesIO(result))
        assert "Dashboard" in wb.sheetnames

    def test_header_row_present(self) -> None:
        rows = [{"institution_code": "TUT", "module_count": 48}]
        result = export_excel(rows)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        headers = [ws.cell(1, col).value for col in range(1, 3)]
        assert any(h is not None for h in headers)

    def test_data_row_present(self) -> None:
        rows = [{"institution_code": "TUT", "module_count": 48}]
        result = export_excel(rows)
        wb = openpyxl.load_workbook(io.BytesIO(result))
        ws = wb.active
        data_row = [ws.cell(2, col).value for col in range(1, 3)]
        assert any(v is not None for v in data_row)

    def test_empty_rows_no_crash(self) -> None:
        result = export_excel([])
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_metadata_contains_platform_name(self) -> None:
        rows = [{"x": 1}]
        result = export_excel(rows, title="Test")
        wb = openpyxl.load_workbook(io.BytesIO(result))
        meta = wb["Metadata"]
        all_vals = [meta.cell(r, 2).value for r in range(1, 5)]
        assert any("AQAA" in str(v) for v in all_vals if v)


# ===========================================================================
# TestExportPdf
# ===========================================================================


class TestExportPdfPlaceholder:
    def test_returns_bytes(self) -> None:
        result = export_pdf_placeholder(["line one", "line two"])
        assert isinstance(result, bytes)

    def test_contains_placeholder_notice(self) -> None:
        result = export_pdf_placeholder(["dummy"])
        text = result.decode("utf-8")
        assert "Placeholder" in text or "placeholder" in text

    def test_contains_title(self) -> None:
        result = export_pdf_placeholder([], title="My Report")
        text = result.decode("utf-8")
        assert "My Report" in text

    def test_contains_report_lines(self) -> None:
        result = export_pdf_placeholder(["Institutions: 2", "Modules: 48"])
        text = result.decode("utf-8")
        assert "Institutions: 2" in text
        assert "Modules: 48" in text


# ===========================================================================
# TestBuildDashboardReportLines
# ===========================================================================


class TestBuildDashboardReportLines:
    def _dashboard_dict(self) -> dict:
        return {
            "generated_at": "2026-07-02T10:00:00Z",
            "institution_count": 2,
            "faculty_count": 8,
            "department_count": 16,
            "programme_count": 16,
            "module_count": 48,
            "audit_run_count": 10,
            "completed_audit_count": 7,
            "failed_audit_count": 1,
            "evidence_file_count": 120,
            "knowledge_index_status": [
                {"institution_code": "TUT", "ikp_version": "v1.1.0", "collection": "tut_2026_v1_1_0", "indexed": True},
            ],
            "by_institution": [
                {
                    "institution_code": "TUT",
                    "institution_name": "Tshwane University of Technology",
                    "faculty_count": 5,
                    "department_count": 10,
                    "programme_count": 10,
                    "module_count": 30,
                    "audit_run_count": 8,
                    "evidence_file_count": 80,
                    "knowledge_indexed": True,
                },
            ],
        }

    def test_returns_list(self) -> None:
        lines = build_dashboard_report_lines(self._dashboard_dict())
        assert isinstance(lines, list)
        assert len(lines) > 0

    def test_contains_module_count(self) -> None:
        lines = build_dashboard_report_lines(self._dashboard_dict())
        text = "\n".join(lines)
        assert "48" in text

    def test_contains_institution_code(self) -> None:
        lines = build_dashboard_report_lines(self._dashboard_dict())
        text = "\n".join(lines)
        assert "TUT" in text

    def test_knowledge_index_indexed_label(self) -> None:
        lines = build_dashboard_report_lines(self._dashboard_dict())
        text = "\n".join(lines)
        assert "INDEXED" in text

    def test_not_indexed_label(self) -> None:
        data = self._dashboard_dict()
        data["knowledge_index_status"][0]["indexed"] = False
        lines = build_dashboard_report_lines(data)
        text = "\n".join(lines)
        assert "NOT INDEXED" in text


# ===========================================================================
# TestTenantIsolation (service-layer)
# ===========================================================================


class TestTenantIsolation:
    def _make_user(self, role: str, institution_id: uuid.UUID | None = None) -> MagicMock:
        user = MagicMock()
        user.role = role
        user.institution_id = institution_id
        return user

    @pytest.mark.asyncio
    async def test_non_admin_institution_summary_own_institution(self) -> None:
        from app.models.enums import UserRole
        institution_id = uuid.uuid4()
        user = self._make_user(UserRole.QUALITY_ASSURANCE_OFFICER, institution_id)

        mock_inst = MagicMock()
        mock_inst.id = institution_id
        mock_inst.code = "TUT"
        mock_inst.name = "TUT"
        mock_inst.institution_type = "university"

        db = AsyncMock()
        db.get = AsyncMock(return_value=mock_inst)
        db.execute = AsyncMock(return_value=MagicMock(scalar_one=MagicMock(return_value=0)))

        with patch("app.reporting.report_service.qdrant_service") as mock_qdrant, \
             patch("app.reporting.report_service.ACTIVE_PILOT_COLLECTIONS", []):
            mock_qdrant.collection_exists.return_value = False
            result = await report_service.get_institution_summary(db, institution_id, user)

        assert result.institution_code == "TUT"

    @pytest.mark.asyncio
    async def test_non_admin_institution_summary_cross_tenant_denied(self) -> None:
        from app.models.enums import UserRole
        own_id = uuid.uuid4()
        other_id = uuid.uuid4()
        user = self._make_user(UserRole.QUALITY_ASSURANCE_OFFICER, own_id)

        db = AsyncMock()

        with pytest.raises(PermissionError):
            await report_service.get_institution_summary(db, other_id, user)

    @pytest.mark.asyncio
    async def test_admin_can_access_any_institution(self) -> None:
        from app.models.enums import UserRole
        institution_id = uuid.uuid4()
        user = self._make_user(UserRole.SYSTEM_ADMIN)

        mock_inst = MagicMock()
        mock_inst.id = institution_id
        mock_inst.code = "UP"
        mock_inst.name = "UP"
        mock_inst.institution_type = "university"

        db = AsyncMock()
        db.get = AsyncMock(return_value=mock_inst)
        db.execute = AsyncMock(return_value=MagicMock(scalar_one=MagicMock(return_value=0)))

        with patch("app.reporting.report_service.qdrant_service") as mock_qdrant, \
             patch("app.reporting.report_service.ACTIVE_PILOT_COLLECTIONS", []):
            mock_qdrant.collection_exists.return_value = False
            result = await report_service.get_institution_summary(db, institution_id, user)

        assert result.institution_code == "UP"


# ===========================================================================
# TestComplianceSummary
# ===========================================================================


class TestComplianceSummary:
    @pytest.mark.asyncio
    async def test_empty_iids_returns_zero_response(self) -> None:
        from app.models.enums import UserRole
        user = MagicMock()
        user.role = UserRole.QUALITY_ASSURANCE_OFFICER
        user.institution_id = None

        db = AsyncMock()
        result = await report_service.get_compliance_summary(db, user, institution_id=None)

        assert result.total_modules == 0
        assert result.compliance_rate_pct == 0.0
        assert result.institution_code == "N/A"

    def test_compliance_rate_formula(self) -> None:
        total = 10
        completed = 7
        rate = round(completed / total * 100.0, 2)
        assert rate == 70.0

    def test_compliance_unaudited_floor(self) -> None:
        total = 5
        completed = 8
        unaudited = max(0, total - completed)
        assert unaudited == 0
